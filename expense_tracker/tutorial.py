#!/usr/bin/python3

from openpyxl import Workbook
wb = Workbook()

# A workbook is always created with at least one worksheet
ws = wb.active

# create new worksheets using the Workbook.create_sheet method
ws1 = wb.create_sheet("test1")  # this will insert it at the end
ws2 = wb.create_sheet("test2", 0)  # insert at first position
ws3 = wb.create_sheet("test3", -1)  # insert at the second to last position

# changing the name of the sheet
ws.title = "New Title"

# change the background color of the tab holding the title
ws.sheet_properties.tabColor = "1072BA"

# once you gave a worksheet a name, you can get it as a key of the workbook
ws3 = wb["New Title"]

# display the names of worksheets of the workbook
for sheet in wb:
    print(sheet.title)

print(ws)
print(ws3)

# Accessing one cell
c = ws['A4']

# values can be directly assigned
ws['A4'] = 4

# Worksheet.cell() method provides access to cells using rows and column notation
d = ws.cell(4, 2, 10)

# Data storage
c.value = 'hello, world'

# Saving to a file
wb.save('tutorial.xlsx')
