#!/usr/bin/python3

from openpyxl import Workbook, load_workbook
import os.path


def greeting():
    initial_choice = input(
        "Welcome to David's Expense Tracker! Type \"new\" to begin a new tracker file, or hit enter to open a previously created file:\n")
    if initial_choice.lower().strip() == "new":
        init()
    else:
        open_prev()


def init():
    wb = Workbook()
    month = input("What month would you like to track the expense?\n")
    dest_filename = month + ".xlsx"
    wb.save(filename=dest_filename)
    edit_tracker(dest_filename)


def open_prev():
    prev_file = input("Which file would you like to open? Just type the name without the extension:\n")
    file_name = prev_file + ".xlsx"
    if os.path.isfile(f"./{file_name}"):
        edit_tracker(file_name)
    else:
        user_input = input("That file does not exist. Would you like to start a new file?\n")
        if user_input.lower().startswith('y'):
            init()
        else:
            open_prev()


def print_rows(sheet):
    for row in sheet.iter_rows(values_only=True):
        print(row)


def edit_tracker(doc_name):
    print("You are in editing mode of file: " + doc_name)
    user_input = input("Enter \"yes\" to continue edit, hit enter to skip:\n")

    workbook = load_workbook(filename=doc_name)
    sheet = workbook.active
    sheet["A1"] = "Day"
    sheet["B1"] = "Money Spent"
    sheet["C1"] = "Description"

    # sheet.delete_rows(idx=2, amount=3)

    if user_input.lower().startswith('y'):
        expense_date = input("What day would you like to add? Enter a number:\n")
        expense_amount = input("Amount spent:\n")
        expense_description = input("What did you spent it on? Enter here:\n")

        new_row = sheet.max_row + 1
        sheet.cell(column=1, row=new_row, value=expense_date)
        sheet.cell(column=2, row=new_row, value=int(expense_amount))
        sheet.cell(column=3, row=new_row, value=expense_description)

    print_rows(sheet)

    workbook.save(filename=doc_name)
    workbook.close()

    display_sum(doc_name)


def display_sum(doc_name):
    print("here")
    workbook = load_workbook(filename=doc_name, data_only=True)
    sheet = workbook.active

    new_row = sheet.max_row + 1
    total_spending = []

    rng = sheet['B2':'B' + str(sheet.max_row)]
    for cell in rng:
        total_spending.append(cell[0].value)

    sheet.cell(column=2, row=new_row, value="total spending is: " + str(sum(total_spending)) + " dollars")

    print_rows(sheet)


greeting()
